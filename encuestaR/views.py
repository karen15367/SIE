from django.shortcuts import redirect, render
from openpyxl import Workbook
from django.db.models import OuterRef, Subquery
from core.models import Encuesta, EncuestaS1, EncuestaS2, EncuestaS3, EncuestaS4, EncuestaS5, EncuestaS3Empresa
from django.http import HttpResponse
from django.db.models import Count
import csv
from core.models import Egresado, Administrador

# Create your views here.


def viewEncuesta(request):

    if request.method == 'POST':

        p = request.POST.get('periodo')

        lista = [
            {'id': 'PERFIL DEL EGRESADO', 'direccion': 'encuestaR/E1'},
            {'id': 'PERTINENCIA Y DISPONIBILIDAD', 'direccion': 'encuestaR/E2'},
            {'id': 'UBICACIÓN LABORAL', 'direccion': 'encuestaR/E3'},
            {'id': 'DESEMPEÑO PROFESIONAL', 'direccion': 'encuestaR/E4'},
            {'id': 'DESAROLLO PROFESIONAL Y SOCIAL', 'direccion': 'encuestaR/E5'},
        ]
        return render(request, 'respuestasListado.html', {
            'anexo': False,
            'inicio': False,
            'resultado': True,
            'listado': lista,
            'periodo': p,

        })
    else:

        listado = Encuesta.objects.values_list('lapso', flat=True).distinct()
        return render(request, 'respuestasListado.html', {
            'anexo': False,
            'inicio': True,
            'resultado': True,
            'listado': listado,
        })


def viewE1(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')
        sF = EncuestaS1.objects.filter(
            sexo=1,
            folioEncuestaS1__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        sM = EncuestaS1.objects.filter(
            sexo=0,
            folioEncuestaS1__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        e1 = EncuestaS1.objects.filter(
            estadoCivil=1,
            folioEncuestaS1__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        e2 = EncuestaS1.objects.filter(
            estadoCivil=2,
            folioEncuestaS1__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        t1 = EncuestaS1.objects.filter(
            titulado=0,
            folioEncuestaS1__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        t2 = EncuestaS1.objects.filter(
            titulado=1,
            folioEncuestaS1__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        # Para p1 y p2, el campo es manejoPaquetes
        p1 = EncuestaS1.objects.filter(
            manejoPaquetes=0,
            folioEncuestaS1__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        p2 = EncuestaS1.objects.filter(
            manejoPaquetes=1,
            folioEncuestaS1__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

    return render(request, 'layouts/E1.html', {
        'anexo': False,
        'subtitle': 'PERFIL DEL EGRESADO',
        'sexo': {'si': sF, 'no': sM},
        'estado': {'si': e1, 'no': e2},
        'titulado': {'si': t1, 'no': t2},
        'paquetes': {'si': p1, 'no': p2},
        'encuesta': p,
        'url': 'encuestaR/exportarE1'
    })


def exportarE1(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')


        data = EncuestaS1.objects.filter(
            folioEncuestaS1__in=Encuesta.objects.filter(
                lapso=p).values('folioEncuesta')
        )

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="encuestaE1_export.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'CURP', 'Sexo', 'Estado civil', 'Titulado', 'Manejo de paquetes'
        ])

        for e in data:
            writer.writerow([
                e.curp,
                'Femenino' if e.sexo == 1 else 'Masculino',
                'Soltero' if e.estadoCivil == 1 else 'Casado',
                'Sí' if e.titulado == 1 else 'No',
                'Sí' if e.manejoPaquetes == 1 else 'No'
            ])
        return response
    


def viewE2(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')

        c1 = EncuestaS2.objects.filter(
            calidadDocentes=1,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        c2 = EncuestaS2.objects.filter(
            calidadDocentes=2,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        c3 = EncuestaS2.objects.filter(
            calidadDocentes=3,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        c4 = EncuestaS2.objects.filter(
            calidadDocentes=4,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        p1 = EncuestaS2.objects.filter(
            planEstudios=1,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        p2 = EncuestaS2.objects.filter(
            planEstudios=2,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        p3 = EncuestaS2.objects.filter(
            planEstudios=3,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        p4 = EncuestaS2.objects.filter(
            planEstudios=4,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        o1 = EncuestaS2.objects.filter(
            oportunidadesProyectos=1,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        o2 = EncuestaS2.objects.filter(
            oportunidadesProyectos=2,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        o3 = EncuestaS2.objects.filter(
            oportunidadesProyectos=3,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        o4 = EncuestaS2.objects.filter(
            oportunidadesProyectos=4,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        e1 = EncuestaS2.objects.filter(
            enfasisInvestigacion=1,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        e2 = EncuestaS2.objects.filter(
            enfasisInvestigacion=2,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        e3 = EncuestaS2.objects.filter(
            enfasisInvestigacion=3,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        e4 = EncuestaS2.objects.filter(
            enfasisInvestigacion=4,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        s1 = EncuestaS2.objects.filter(
            satisfaccionCondiciones=1,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        s2 = EncuestaS2.objects.filter(
            satisfaccionCondiciones=2,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        s3 = EncuestaS2.objects.filter(
            satisfaccionCondiciones=3,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        s4 = EncuestaS2.objects.filter(
            satisfaccionCondiciones=4,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        x1 = EncuestaS2.objects.filter(
            experienciaResidencia=1,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        x2 = EncuestaS2.objects.filter(
            experienciaResidencia=2,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        x3 = EncuestaS2.objects.filter(
            experienciaResidencia=3,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        x4 = EncuestaS2.objects.filter(
            experienciaResidencia=4,
            folioEncuestaS2__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

    return render(request, 'layouts/E2.html', {
        'anexo': False,
        'subtitle': 'PERTINENCIA Y DISPONIBILIDAD DE MEDIOS Y RECURSOS PARA EL APRENDIZAJE',
        'calidad': {'uno': c1, 'dos': c2, 'tres': c3, 'cuatro': c4},
        'plan': {'uno': p1, 'dos': p2, 'tres': p3, 'cuatro': p4},
        'oportunidad': {'uno': o1, 'dos': o2, 'tres': o3, 'cuatro': o4},
        'enfasis': {'uno': e1, 'dos': e2, 'tres': e3, 'cuatro': e4},
        'satisfaccion': {'uno': s1, 'dos': s2, 'tres': s3, 'cuatro': s4},
        'experiencia': {'uno': x1, 'dos': x2, 'tres': x3, 'cuatro': x4},
        'encuesta': p,
        'url': 'encuestaR/exportarE2'

    })


def exportarE2(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')

    data = EncuestaS2.objects.select_related('folioEncuesta__curp').filter(
        folioEncuestaS2__in=Encuesta.objects.filter(
            lapso=p).values('folioEncuesta')
    )

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="encuestaE2_export.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'CURP', 'Calidad docentes', 'Plan estudios',
        'Oportunidad en proyectos', 'Énfasis en investigación',
        'Satisfacción condiciones de estudio', 'Experiencia residencia'
    ])

    for e in data:
        writer.writerow([
            e.folioEncuesta.curp.curp,
            e.get_calidadDocentes_display() if e.calidadDocentes else '',
            e.get_planEstudios_display() if e.planEstudios else '',
            e.get_oportunidadesProyectos_display() if e.oportunidadesProyectos else '',
            e.get_enfasisInvestigacion_display() if e.enfasisInvestigacion else '',
            e.get_satisfaccionCondiciones_display() if e.satisfaccionCondiciones else '',
            e.get_experienciaResidencia_display() if e.experienciaResidencia else ''
        ])

    return response


def viewE3(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')
        # answers = EncuestaS3.objects.all()

        answers = EncuestaS3.objects.filter(
            folioEncuestaS3__in=Encuesta.objects.filter(
                lapso=p).values('folioEncuesta')
        )

        a1 = EncuestaS3.objects.filter(
            actividad=1,
            folioEncuestaS3__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        a2 = EncuestaS3.objects.filter(
            actividad=2,
            folioEncuestaS3__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        a3 = EncuestaS3.objects.filter(
            actividad=3,
            folioEncuestaS3__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        a4 = EncuestaS3.objects.filter(
            actividad=4,
            folioEncuestaS3__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        # answersE = EncuestaS3Empresa.objects.all()

        answersE = EncuestaS3Empresa.objects.filter(
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=p).values('folioEncuesta')
        )

    # * Estudian
        b1 = sum(1 for a in answers if (a.actividad ==
                                        3 or a.actividad == 2) and a.tipoEstudio == 1)
        b2 = sum(1 for a in answers if (a.actividad ==
                                        3 or a.actividad == 2) and a.tipoEstudio == 2)
        b3 = sum(1 for a in answers if (a.actividad ==
                                        3 or a.actividad == 2) and a.tipoEstudio == 3)
        b4 = sum(1 for a in answers if (a.actividad ==
                                        3 or a.actividad == 2) and a.tipoEstudio == 4)
        b5 = sum(1 for a in answers if (a.actividad ==
                                        3 or a.actividad == 2) and a.tipoEstudio == 5)
# * trabajan
        t1 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.tipoEstudio == 1)
        t2 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.tipoEstudio == 2)
        t3 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.tipoEstudio == 3)
        t4 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.tipoEstudio == 4)

        m1 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.medioEmpleo == 1)
        m2 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.medioEmpleo == 2)
        m3 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.medioEmpleo == 3)
        m4 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.medioEmpleo == 4)
        m5 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.medioEmpleo == 5)

        r1 = sum(1 for a in answers if (a.actividad == 1 or a.actividad == 3)
                 and a.requisitosContratacion == 1)
        r2 = sum(1 for a in answers if (a.actividad == 1 or a.actividad == 3)
                 and a.requisitosContratacion == 2)
        r3 = sum(1 for a in answers if (a.actividad == 1 or a.actividad == 3)
                 and a.requisitosContratacion == 3)
        r4 = sum(1 for a in answers if (a.actividad == 1 or a.actividad == 3)
                 and a.requisitosContratacion == 4)
        r5 = sum(1 for a in answers if (a.actividad == 1 or a.actividad == 3)
                 and a.requisitosContratacion == 5)
        r6 = sum(1 for a in answers if (a.actividad == 1 or a.actividad == 3)
                 and a.requisitosContratacion == 6)
        r7 = sum(1 for a in answers if (a.actividad == 1 or a.actividad == 3)
                 and a.requisitosContratacion == 7)

        i1 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.idiomaUtiliza == 1)
        i2 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.idiomaUtiliza == 2)
        i3 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.idiomaUtiliza == 3)
        i4 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.idiomaUtiliza == 4)
        i5 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.idiomaUtiliza == 5)

        e1 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.antiguedad == 1)
        e2 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.antiguedad == 2)
        e3 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.antiguedad == 3)
        e4 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.antiguedad == 4)
        e5 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.antiguedad == 5)

        s1 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.ingreso == 1)
        s2 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.ingreso == 2)
        s3 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.ingreso == 3)
        s4 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.ingreso == 4)

        n1 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.nivelJerarquico == 1)
        n2 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.nivelJerarquico == 2)
        n3 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.nivelJerarquico == 3)
        n4 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.nivelJerarquico == 4)
        n5 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.nivelJerarquico == 5)
        n6 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.nivelJerarquico == 6)

        c1 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.condicionTrabajo == 1)
        c2 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.condicionTrabajo == 2)
        c3 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.condicionTrabajo == 3)
        c4 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.condicionTrabajo == 4)

        r1 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.nivelJerarquico == 1)
        r2 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.nivelJerarquico == 2)
        r3 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.nivelJerarquico == 3)
        r4 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.nivelJerarquico == 4)
        r5 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.nivelJerarquico == 5)
        r6 = sum(1 for a in answers if (a.actividad ==
                                        1 or a.actividad == 3) and a.nivelJerarquico == 6)

        u1 = EncuestaS3Empresa.objects.filter(
            tipoOrganismo=1,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        u2 = EncuestaS3Empresa.objects.filter(
            tipoOrganismo=2,  # Asumiendo que hay diferentes tipos de organismo
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        u3 = EncuestaS3Empresa.objects.filter(
            tipoOrganismo=3,  # Asumiendo que hay diferentes tipos de organismo
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        q1 = EncuestaS3Empresa.objects.filter(
            sectorPrimario=1,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        q2 = EncuestaS3Empresa.objects.filter(
            sectorPrimario=2,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        q3 = EncuestaS3Empresa.objects.filter(
            sectorPrimario=3,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        q4 = EncuestaS3Empresa.objects.filter(
            sectorPrimario=4,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        h1 = EncuestaS3Empresa.objects.filter(
            sectorSecundario=1,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        h2 = EncuestaS3Empresa.objects.filter(
            sectorSecundario=2,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        h3 = EncuestaS3Empresa.objects.filter(
            sectorSecundario=3,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        h4 = EncuestaS3Empresa.objects.filter(
            sectorSecundario=4,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        y1 = EncuestaS3Empresa.objects.filter(
            sectorTerciario=1,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        y2 = EncuestaS3Empresa.objects.filter(
            sectorTerciario=2,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        y3 = EncuestaS3Empresa.objects.filter(
            sectorTerciario=3,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

        y4 = EncuestaS3Empresa.objects.filter(
            sectorTerciario=4,
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
        ).count()

    return render(request, 'layouts/E3.html', {
        'anexo': False,
        'subtitle': 'UBICACIÓN LABORAL DE LOS EGRESADOS',
        'actividad': {'uno': a1, 'dos': a2, 'tres': a3, 'cuatro': a4},
        'estudio': {'uno': b1, 'dos': b2, 'tres': b3, 'cuatro': b4, 'cinco': b5},
        'trabajo': {'uno': t1, 'dos': t2, 'tres': t3, 'cuatro': t4},
        'medio': {'uno': m1, 'dos': m2, 'tres': m3, 'cuatro': m4, 'cinco': m5},
        'requisitos': {'uno': r1, 'dos': r2, 'tres': r3, 'cuatro': r4, 'cinco': r5, 'seis': r6, 'siete': r7},
        'idioma': {'uno': i1, 'dos': i2, 'tres': i3, 'cuatro': i4, 'cinco': i5},
        'empleo': {'uno': e1, 'dos': e2, 'tres': e3, 'cuatro': e4, 'cinco': e5},
        'salario': {'uno': s1, 'dos': s2, 'tres': s3, 'cuatro': s4, },
        'nivel': {'uno': n1, 'dos': n2, 'tres': n3, 'cuatro': n4, 'cinco': n5, 'seis': n6},
        'condicion': {'uno': c1, 'dos': c2, 'tres': c3, 'cuatro': c4, },
        'relacion': {'uno': r1, 'dos': r2, 'tres': r3, 'cuatro': r4, 'cinco': r5, 'seis': r6},
        'empresa': {'uno': u1, 'dos': u2, 'tres': u3, },
        'sectorp': {'uno': q1, 'dos': q2, 'tres': q3, 'cuatro': q4, },
        'sectors': {'uno': h1, 'dos': h2, 'tres': h3, 'cuatro': h4, },
        'sectort': {'uno': y1, 'dos': y2, 'tres': y3, 'cuatro': y4, },
        'encuesta': p,
        'url': 'encuestaR/exportarE3'
    })


def exportarE3(request):

    if request.method == 'POST':
        p = request.POST.get('periodo')


        encuestas = EncuestaS3.objects.select_related('folioEncuesta').filter(
            folioEncuestaS3__in=Encuesta.objects.filter(lapso=p).values('folioEncuesta')
        )

        empresas = EncuestaS3Empresa.objects.select_related('folioEncuesta').filter(
            folioEncuestaS3Empresa__in=Encuesta.objects.filter(lapso=p).values('folioEncuesta')
        )

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="encuestaE3_export.csv"'
        writer = csv.writer(response)

        # Tabla 1: EncuestaS3 (egresado)
        writer.writerow(['SECCIÓN 1: DATOS DEL EGRESADO'])
        writer.writerow([
            'Folio Encuesta',
            'Actividad actual',
            'Tipo de estudio',
            'Medio de empleo',
            'Requisitos de contratación',
            'Idioma que utiliza',
            'Antigüedad en el empleo',
            'Ingreso diario',
            'Nivel jerárquico',
            'Condición de trabajo',
            'Relación con formación'
        ])

        for e in encuestas:
            writer.writerow([
                e.folioEncuesta_id,
                e.get_actividad_display() if e.actividad else '',
                e.get_tipoEstudio_display() if e.tipoEstudio else '',
                e.get_medioEmpleo_display() if e.medioEmpleo else '',
                e.get_requisitosContratacion_display() if e.requisitosContratacion else '',
                e.get_idiomaUtiliza_display() if e.idiomaUtiliza else '',
                e.get_antiguedad_display() if e.antiguedad else '',
                e.get_ingreso_display() if e.ingreso else '',
                e.get_nivelJerarquico_display() if e.nivelJerarquico else '',
                e.get_condicionTrabajo_display() if e.condicionTrabajo else '',
                e.get_relacionTrabajo_display() if e.relacionTrabajo else ''
            ])

        writer.writerow([])  # línea en blanco

        # Tabla 2: EncuestaS3Empresa
        writer.writerow(['SECCIÓN 2: DATOS DE LA EMPRESA U ORGANIZACIÓN'])
        writer.writerow([
            'Folio Encuesta',
            'Tipo de organismo',
            'Tamaño de empresa',
            'Sector primario',
            'Sector secundario',
            'Sector terciario'
        ])

        for e in empresas:
            writer.writerow([
                e.folioEncuesta_id,
                e.get_tipoOrganismo_display() if e.tipoOrganismo else '',
                e.get_tamanoEmpresa_display() if e.tamanoEmpresa else '',
                e.get_sectorPrimario_display() if e.sectorPrimario else '',
                e.get_sectorSecundario_display() if e.sectorSecundario else '',
                e.get_sectorTerciario_display() if e.sectorTerciario else '',
            ])

    return response


def viewE4(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')

        def contar(campo, choices, p):
            filtros_modificados = {}
            for key, value in p.items():
                if key == 'lapso':
                    filtros_modificados['folioEncuesta__lapso'] = value
                else:
                    filtros_modificados[key] = value

            data = EncuestaS4.objects.filter(
                **filtros_modificados).values(campo).annotate(total=Count(campo))

            # Crear diccionario con las choices
            resultado = {}
            choices_dict = dict(choices)  # Convertir choices a diccionario

            for choice_value, choice_label in choices:
                total = sum(d['total']
                            for d in data if d[campo] == choice_value)
                resultado[choice_label] = total

            return resultado

        context = {
            'areaCampoEstudio': contar('areaCampoEstudio', EncuestaS4.VALORACION_CHOICES, {'lapso': f'{p}'}),
            'titulacion': contar('titulacion', EncuestaS4.VALORACION_CHOICES, {'lapso': f'{p}'}),
            'experienciaLaboral': contar('experienciaLaboral', EncuestaS4.VALORACION_CHOICES, {'lapso': f'{p}'}),
            'competenciaLaboral': contar('competenciaLaboral', EncuestaS4.VALORACION_CHOICES, {'lapso': f'{p}'}),
            'posicionamientoInstitucion': contar('posicionamientoInstitucion', EncuestaS4.VALORACION_CHOICES, {'lapso': f'{p}'}),
            'conocimientoIdiomas': contar('conocimientoIdiomas', EncuestaS4.VALORACION_CHOICES, {'lapso': f'{p}'}),
            'recomendacionesReferencias': contar('recomendacionesReferencias', EncuestaS4.VALORACION_CHOICES, {'lapso': f'{p}'}),
            'personalidadActitudes': contar('personalidadActitudes', EncuestaS4.VALORACION_CHOICES, {'lapso': f'{p}'}),
            'capacidadLiderazgo': contar('capacidadLiderazgo', EncuestaS4.VALORACION_CHOICES, {'lapso': f'{p}'}),
            'otrosAspectos': contar('otrosAspectos', EncuestaS4.VALORACION_CHOICES, {'lapso': f'{p}'}),
            'encuesta': p,
            'url': 'encuestaR/exportarE4'
        }

    return render(request, 'layouts/E4.html', context)


def exportarE4(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')

        data = EncuestaS4.objects.select_related('folioEncuesta').filter(
    folioEncuestaS4__in=Encuesta.objects.filter(lapso=p).values('folioEncuesta')
)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="encuestaE4_export.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Folio Encuesta',
            'Área o campo de estudio',
            'Titulación',
            'Experiencia laboral',
            'Competencia laboral',
            'Posicionamiento institución',
            'Conocimiento idiomas',
            'Recomendaciones/referencias',
            'Personalidad/actitudes',
            'Capacidad liderazgo',
            'Otros aspectos'
        ])

        for e in data:
            writer.writerow([
                e.folioEncuesta_id,
                e.get_areaCampoEstudio_display() if e.areaCampoEstudio else '',
                e.get_titulacion_display() if e.titulacion else '',
                e.get_experienciaLaboral_display() if e.experienciaLaboral else '',
                e.get_competenciaLaboral_display() if e.competenciaLaboral else '',
                e.get_posicionamientoInstitucion_display() if e.posicionamientoInstitucion else '',
                e.get_conocimientoIdiomas_display() if e.conocimientoIdiomas else '',
                e.get_recomendacionesReferencias_display() if e.recomendacionesReferencias else '',
                e.get_personalidadActitudes_display() if e.personalidadActitudes else '',
                e.get_capacidadLiderazgo_display() if e.capacidadLiderazgo else '',
                e.get_otrosAspectos_display() if e.otrosAspectos else '',
            ])

    return response


def viewE5(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')

        def contar(campo, p):
            filtros_modificados = {}
            for key, value in p.items():
                if key == 'lapso':
                    # Ajustar según tu relación
                    filtros_modificados['folioEncuesta__lapso'] = value
                else:
                    filtros_modificados[key] = value

            data = EncuestaS5.objects.filter(
                **filtros_modificados).values(campo).annotate(total=Count(campo))
            return {'Sí': sum(d['total'] for d in data if d[campo] == 1), 'No': sum(d['total'] for d in data if d[campo] == 0)}

        context = {
            'cursosActualizacion': contar('cursosActualizacion', {'lapso': f'{p}'}),
            'tomarPosgrado': contar('tomarPosgrado', {'lapso': f'{p}'}),
            'orgSociales': contar('perteneceOrganizacionesSociales', {'lapso': f'{p}'}),
            'orgProfesionales': contar('perteneceOrganismosProfesionistas', {'lapso': f'{p}'}),
            'asociacionEgresados': contar('perteneceAsociacionEgresados', {'lapso': f'{p}'}),
            'encuesta': p,
            'url': 'encuestaR/exportarE5'
        }
    return render(request, 'layouts/E5.html', context)


def exportarE5(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')
        answer = EncuestaS5.objects.filter(
    folioEncuestaS5__in=Encuesta.objects.filter(lapso=p).values('folioEncuesta')
)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="E5.csv"'
        writer = csv.writer(response)
        writer.writerow(['Folio', 'Cursos Actualización', 'Desea Posgrado',
                        'Org. Sociales', 'Org. Profesionales', 'Asoc. Egresados'])
        for e in answer:
            writer.writerow([
                e.folioEncuesta_id,
                e.cursosActualizacion,
                e.tomarPosgrado,
                e.perteneceOrganizacionesSociales,
                e.perteneceOrganismosProfesionistas,
                e.perteneceAsociacionEgresados
            ])
    return response


def formulario_export_encuesta(request):
    carreras = [
        "general",
        "Ing. Sistemas Computacionales",
        "Ing. Industrial",
        "Ing. Mecatrónica",
        "Ing. Eléctrica",
        "Ing. Electrónica",
        "Ing. Mecánica",
        "Ing. Renovables",
        "Ing. Gestión Empresarial",
        "Lic. Administración"
    ]
    return render(request, 'ExportarEncuestaUsuario.html', {'carreras': carreras})

def exportar_encuesta_por_carrera(request):
    carrera = request.GET.get('carrera')
    if not carrera:
        return HttpResponse("Carrera no seleccionada", status=400)

    rfc_admin = request.session.get('usuario_id')
    if not rfc_admin:
        return HttpResponse("Sesión no válida. Inicia sesión nuevamente.", status=401)

    try:
        admin = Administrador.objects.get(rfc=rfc_admin)
    except Administrador.DoesNotExist:
        return HttpResponse("Administrador no encontrado.", status=404)

    if admin.carrera.lower() != "general" and admin.carrera.lower() != carrera.lower():
        return HttpResponse("No tienes permiso para exportar esta carrera.", status=403)

    if carrera == "general":
        egresados_curps = list(Egresado.objects.values_list('curp', flat=True))
    else:
        egresados_curps = list(Egresado.objects.filter(carrera=carrera).values_list('curp', flat=True))

    ultimos = Encuesta.objects.filter(
        curp=OuterRef('folioEncuesta__curp')
    ).order_by('-lapso').values('folioEncuesta')[:1]

    wb = Workbook()
    sheet_map = {
        'Encuesta S1': (EncuestaS1, wb.active),
        'Encuesta S2': (EncuestaS2, wb.create_sheet('Encuesta S2')),
        'Encuesta S3': (EncuestaS3, wb.create_sheet('Encuesta S3')),
        'Empresa': (EncuestaS3Empresa, wb.create_sheet('Empresa')),
        'Encuesta S4': (EncuestaS4, wb.create_sheet('Encuesta S4')),
        'Encuesta S5': (EncuestaS5, wb.create_sheet('Encuesta S5')),
    }
    sheet_map['Encuesta S1'][1].title = 'Encuesta S1'

    def get_display(obj, field):
        val = getattr(obj, field.name)
        if field.choices:
            return dict(field.flatchoices).get(val, val)
        return val

    for sheet_name, (ModelClass, sheet) in sheet_map.items():
        queryset = ModelClass.objects.filter(
            folioEncuesta_id=Subquery(ultimos),
            folioEncuesta__curp__curp__in=egresados_curps
        ).select_related('folioEncuesta__curp')

        field_names = [f.verbose_name.title() if f.verbose_name else f.name for f in ModelClass._meta.fields if f.name != 'id' and not f.name.startswith('folioEncuesta')]
        sheet.append(['CURP', 'Carrera', 'Lapso'] + field_names)

        if queryset.exists():
            for obj in queryset:
                fila = [
                    obj.folioEncuesta.curp.curp,
                    obj.folioEncuesta.curp.carrera,
                    obj.folioEncuesta.lapso
                ]
                for field in ModelClass._meta.fields:
                    if field.name != 'id' and not field.name.startswith('folioEncuesta'):
                        fila.append(get_display(obj, field))
                sheet.append(fila)
        else:
            # Agrega una fila vacía si no hay datos
            empty_row = ['' for _ in range(3 + len(field_names))]
            sheet.append(empty_row)

    nombre_archivo = 'encuesta_general.xlsx' if carrera == "general" else f'encuesta_{carrera}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


def get_display(obj, field):
    val = getattr(obj, field.name)

    # Valor nulo → celda vacía
    if val is None:
        return ''

    # Usa display si tiene choices
    if field.choices:
        return dict(field.flatchoices).get(val, val)

    # Por compatibilidad: campos típicos 1/0 sin choices definidos
    if isinstance(val, bool) or field.get_internal_type() == 'BooleanField':
        return 'Sí' if val else 'No'
    
    return val


def exportar_encuesta_por_lapso(request):
    lapso = request.GET.get('lapso')
    if not lapso:
        return HttpResponse("Lapso no especificado", status=400)

    rfc_admin = request.session.get('usuario_id')
    if not rfc_admin:
        return HttpResponse("Sesión no válida. Inicia sesión nuevamente.", status=401)

    try:
        admin = Administrador.objects.get(rfc=rfc_admin)
    except Administrador.DoesNotExist:
        return HttpResponse("Administrador no encontrado.", status=404)

    # CURPs válidos según la carrera del admin
    if admin.carrera.lower() == "general":
        egresados_curps = list(Egresado.objects.values_list('curp', flat=True))
    else:
        egresados_curps = list(Egresado.objects.filter(carrera=admin.carrera).values_list('curp', flat=True))

    # Encuestas por lapso
    encuestas_lapso = Encuesta.objects.filter(lapso=lapso, curp__in=egresados_curps).values_list('folioEncuesta', flat=True)

    from openpyxl import Workbook
    wb = Workbook()

    # Mapeo de modelos y hojas
    sheet_map = {
        'Encuesta S1': (EncuestaS1, wb.active),
        'Encuesta S2': (EncuestaS2, wb.create_sheet('Encuesta S2')),
        'Encuesta S3': (EncuestaS3, wb.create_sheet('Encuesta S3')),
        'Empresa': (EncuestaS3Empresa, wb.create_sheet('Empresa')),
        'Encuesta S4': (EncuestaS4, wb.create_sheet('Encuesta S4')),
        'Encuesta S5': (EncuestaS5, wb.create_sheet('Encuesta S5')),
    }
    sheet_map['Encuesta S1'][1].title = 'Encuesta S1'

    def get_display(obj, field):
        val = getattr(obj, field.name)
        if val is None:
            return ''
        if field.choices:
            return dict(field.flatchoices).get(val, val)
        if isinstance(val, bool) or field.get_internal_type() == 'BooleanField':
            return 'Sí' if val else 'No'
        return val

    total_registros = 0  # Para validar si hay datos

    for sheet_name, (ModelClass, sheet) in sheet_map.items():
        queryset = ModelClass.objects.filter(
            folioEncuesta_id__in=encuestas_lapso
        ).select_related('folioEncuesta__curp')

        # Encabezados
        field_names = [
            f.verbose_name.title() if f.verbose_name else f.name
            for f in ModelClass._meta.fields
            if f.name != 'id' and not f.name.startswith('folioEncuesta')
        ]
        sheet.append(['CURP', 'Carrera', 'Lapso'] + field_names)

        for obj in queryset:
            total_registros += 1
            fila = [
                obj.folioEncuesta.curp.curp,
                obj.folioEncuesta.curp.carrera,
                obj.folioEncuesta.lapso
            ]
            for field in ModelClass._meta.fields:
                if field.name != 'id' and not field.name.startswith('folioEncuesta'):
                    fila.append(get_display(obj, field))
            sheet.append(fila)

        if queryset.count() == 0:
            # Hoja vacía con fila en blanco
            sheet.append(['' for _ in range(3 + len(field_names))])

    if total_registros == 0:
        return HttpResponse("No hay datos para exportar en el lapso seleccionado.", status=204)

    # Nombre limpio del archivo
    safe_lapso = lapso.replace("/", "-").replace(" ", "_")
    nombre_archivo = f"encuestas_Lapso_{safe_lapso}.xlsx"

    # Respuesta con archivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response
