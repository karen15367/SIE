from django.shortcuts import render
from openpyxl import Workbook
from django.db.models import Count, Q
from core.models import Administrador, AnexoS1, AnexoS2, AnexoS3, AnexoS4, Egresado, Encuesta
from .forms import FiltroEncuestaForm
from django.http import HttpResponse
import csv
from django.db.models import Max, Subquery, OuterRef

# Create your views here.


def viewAnexos(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')

        try:
            lista = [
                {'id': 'DATOS GENERALES', 'direccion': 'anexoR/A1'},
                {'id': 'SITUACIÓN LABORAL', 'direccion': 'anexoR/A2'},
                {'id': 'PLAN DE ESTUDIOS / INSTITUCIÓN ', 'direccion': 'anexoR/A3'},
                {'id': 'DESEMPEÑO LABORAL', 'direccion': 'anexoR/A4'},
            ]

            return render(request, 'respuestasListado.html', {
                'anexo': True,
                'inicio': False,
                'resultado': True,
                'listado': lista,
                'periodo': p,
            })
        except ZeroDivisionError as e:
            print(e)
            return render(request, 'respuestasListado.html', {
                'anexo': True,
                'inicio': False,
                'listado': False
            })
    else:
        lista = Encuesta.objects.values_list('lapso', flat=True).distinct()
        resultado = False
        if lista:
            resultado = True

        return render(request, 'respuestasListado.html', {
            'anexo': True,
            'inicio': 'si',
            'resultado': resultado,
            'listado': lista,

        })


def viewA1(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')

        titulo = AnexoS1.objects.filter(
            titulado=1,
            folioAnexoS1__in=Encuesta.objects.filter(
                lapso=f"{p}").values('anexos1')
        ).count()

        sinTitulo = AnexoS1.objects.filter(
            titulado=0,
            folioAnexoS1__in=Encuesta.objects.filter(
                lapso=f"{p}").values('anexos1')
        ).count()

    return render(request, 'layouts/A1.html', {
        'anexo': True,
            'subtitle': 'INFORMACIÓN DEL EGRESADO',
            'titulo': {'si': titulo, 'no': sinTitulo},
            'encuesta': p,
            'url': 'encuestaR/exportarE1'
    })


def exportarA1(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')
        data = AnexoS1.objects.filter(
            folioEncuesta__in=Encuesta.objects.filter(
                lapso=p).values('folioEncuesta')
        ).select_related('folioEncuesta__curp').all()

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="anexoA1_export.csv"'

        writer = csv.writer(response)
        writer.writerow(['CURP', 'Nombre', 'Correo', 'Sexo',
                        'Carrera', 'Fecha Egreso', 'Titulado'])

        for a in data:
            egresado = a.folioEncuesta.curp
            writer.writerow([
                egresado.curp,
                a.nombreCompleto or '',
                a.correo or '',
                'Femenino' if egresado.sexo == 1 else 'Masculino',
                egresado.carrera or '',
                a.fechaEgreso or '',
                'Sí' if a.titulado == 1 else 'No'
            ])

    return response


def viewA2(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')
        answer = AnexoS2.objects.all()

        trabaja = noTrabaja = relacion = noRelacion = 0

        a1 = a2 = a3 = a4 = 0

        t1 = t2 = t3 = 0

        s1 = s2 = s3 = s4 = 0

        p1 = p2 = p3 = p4 = p5 = p6 = p7 = 0

        d1 = d2 = d3 = d4 = d5 = d6 = d7 = 0

        m1 = m2 = m3 = m4 = m5 = 0

        g1 = g2 = g3 = g4 = 0

        for a in answer:
            if a.trabaja == 1:
                # * El egresado trabaja
                trabaja += 1
                # * Relación entre su trabajo y su carrera
                if a.relacionTrabajoCarrera == 1:
                    relacion += 1
                    # * tiempo en coseguir empleo relacionado con la carrera
                    if a.tiempoTrabajoRelacionado == 1:
                        t1 += 1
                    elif a.tiempoTrabajoRelacionado == 2:
                        t2 += 1
                    elif a.tiempoTrabajoRelacionado == 3:
                        t3 += 1
                else:
                    #! el trabajo no tiene relación con su carrera
                    noRelacion += 1

            # * Tiempo en el que están trabajando
        a1 = AnexoS2.objects.filter(
                actividad=1,
                folioEncuestaS3__in=Encuesta.objects.filter(
                lapso=f"{p}").values('folioEncuesta')
                ).count()

        a2 = AnexoS2.objects.filter(
                actividad=2,
                folioEncuestaS3__in=Encuesta.objects.filter(
                    lapso=f"{p}").values('folioEncuesta')
            ).count()

        a3 = AnexoS2.objects.filter(
                actividad=3,
                folioEncuestaS3__in=Encuesta.objects.filter(
                    lapso=f"{p}").values('folioEncuesta')
            ).count()

        a4 = AnexoS2.objects.filter(
                actividad=4,
                folioEncuestaS3__in=Encuesta.objects.filter(
                    lapso=f"{p}").values('folioEncuesta')
            ).count()

           # * sector en el que trabaja
        if a.sector == 1:
                s1 += 1
        elif a.sector == 2:
                s2 += 1
        elif a.sector == 3:
                s3 += 1
        else:
                s4 += 1

            # * Puesto en su trabajo
        if a.rol == 1:
                p1 += 1
        elif a.rol == 2:
                p2 += 1
        elif a.rol == 3:
                p3 += 1
        elif a.rol == 4:
                p4 += 1
        elif a.rol == 5:
                p5 += 1
        elif a.rol == 6:
                p6 += 1
        else:
                p7 += 1

            # * medio de conseguir trabajo
        if a.medioTrabajo == 1:
                m1 += 1
        elif a.medioTrabajo == 2:
                m2 += 1
        elif a.medioTrabajo == 3:
                m3 += 1
        elif a.medioTrabajo == 4:
                m4 += 1
        else:
                m5 += 1

            # * grado de satisfacción
        if a.satisfaccion == 1:
                g1 += 1
        elif a.satisfaccion == 2:
                g2 += 1
        elif a.satisfaccion == 3:
                g3 += 1
        else:
                g4 += 1

        #! el egresado no trabaja
    else:
            noTrabaja += 1

    return render(request, 'layouts/A2.html', {
        'anexo': True,
        'subtitle': 'SITUACIÓN LABORAL',
        'trabaja': {'si': trabaja, 'no': noTrabaja},  # *si
        'relacion': {'si': relacion, 'no': noRelacion},  # * si
        'antiguedad': {'uno': a1, 'dos': a2, 'tres': a3, 'cuatro': a4},  # *ya
        'trabajoT': {'uno': t1, 'dos': t2, 'tres': t3, },  # * al toke
        'sector': {'uno': s1, 'dos': s2, 'tres': s3, 'cuatro': s4},  # *yes u do
        'puesto': {'uno': p1, 'dos': p2, 'tres': p3, 'cuatro': p4, 'cinco': p5, 'seis': p6, 'siete': p7},
        'area': {'uno': d1, 'dos': d2, 'tres': d3, 'cuatro': d4, 'cinco': d5, 'seis': d6, 'siete': d7},
        'medio': {'uno': m1, 'dos': m2, 'tres': m3, 'cuatro': m4, 'cinco': m5, },  # *
        'satisfaccion': {'uno': g1, 'dos': g2, 'tres': g3, 'cuatro': g4, },
        'encuesta': p,
        'url': 'encuestaR/exportarA2'
    })


def exportarA2(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')
        data = AnexoS2.objects.filter(
            folioEncuesta__in=Encuesta.objects.filter(lapso=p).values('folioEncuesta')
        ).select_related('folioEncuesta__curp').all()

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="anexoA2_export.csv"'

        writer = csv.writer(response)
        writer.writerow(['CURP', 'Titulado', 'Razón No Título', 'Otra Razón'])

        for a in data:
            egresado = a.folioEncuesta.curp
            writer.writerow([
                egresado.curp,
                'Sí' if a.titulado == 1 else 'No',
                a.get_razonNoTitulo_display() if a.razonNoTitulo else '',
                a.razonNoTituloOtra or ''
            ])

    return response


def viewA3(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')
        answer = AnexoS3.objects.all()

        c1 = c2 = c3 = c4 = 0

        g1 = g2 = g3 = g4 = 0

        r1 = r2 = r3 = r4 = r5 = 0

        a1 = a2 = 0

        i1 = i2 = 0

        p1 = p2 = p3 = p4 = p5 = p6 = p7 = 0

        for a in answer:
            if a.competencias == 1:
                c1 += 1
            elif a.competencias == 2:
                c2 += 1
            elif a.competencias == 3:
                c3 += 1
            else:
                c4 += 1

            # * grado de satisfacción
            if a.satisfaccion == 1:
                g1 += 1
            elif a.satisfaccion == 2:
                g2 += 1
            elif a.satisfaccion == 3:
                g3 += 1
            else:
                g4 += 1

            if a.educativo == 1:
                r1 += 1
            elif a.educativo == 2:
                r2 += 1
            elif a.educativo == 3:
                r3 += 1
            elif a.educativo == 4:
                r4 += 1
            else:
                r5 += 1

            if a.contacto == 1:
                a1 += 1
            else:
                a2 += 1

            if a.participar == 1:
                i1 += 1
                if a.aporte == 1:
                    p1 += 1
                elif a.aporte == 2:
                    p2 += 1
                elif a.aporte == 3:
                    p3 += 1
                elif a.aporte == 4:
                    p4 += 1
                elif a.aporte == 5:
                    p5 += 1
                elif a.aporte == 6:
                    p6 += 1
                else:
                    p7 += 1

            else:
                i2 += 1

    return render(request, 'layouts/A3.html', {
        'anexo': True,
        'subtitle': 'PLAN DE ESTUDIOS / INSTITUCIÓN',
        'competencias': {'uno': c1, 'dos': c2, 'tres': c3, 'cuatro': c4},
        'satisfaccion': {'uno': g1, 'dos': g2, 'tres': g3, 'cuatro': g4, },
        'reforzar': {'uno': r1, 'dos': r2, 'tres': r3, 'cuatro': r4, 'cinco': r5},
        'contacto': {'si': a1, 'no': a2},
        'participar': {'si': i1, 'no': i2},
        'aporte': {'uno': p1, 'dos': p2, 'tres': p3, 'cuatro': p4, 'cinco': p5, 'seis': p6, 'siete': p7},
        'encuesta': p,
        'url': 'encuestaR/exportarA3'

    })


def exportarA3(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')
        data = AnexoS3.objects.filter(
            folioEncuesta__in=Encuesta.objects.filter(lapso=p).values('folioEncuesta')
        ).select_related('folioEncuesta__curp').all()

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="anexoA3_export.csv"'

        writer = csv.writer(response)
        writer.writerow(['CURP', 'Nombre completo', 'Teléfono',
                        'Correo', 'Redes sociales'])

        for a in data:
            egresado = a.folioEncuesta.curp
            writer.writerow([
                egresado.curp,
                a.nombreCompleto or '',
                a.telefono or '',
                a.correo or '',
                a.redesSociales or ''
            ])

    return response


def viewA4(request):

    if request.method == 'POST':
        p = request.POST.get('periodo')

        answer = AnexoS4.objects.all()

        h1 = h2 = h3 = h4 = h5 = 0

        c1 = c2 = 0

        d1 = d2 = d3 = d4 = d5 = 0

        r1 = r2 = 0

        z1 = z2 = 0

        s1 = s2 = s3 = s4 = s5 = 0

        i1 = i2 = i3 = i4 = i5 = 0

        p1 = p2 = 0

        w1 = w2 = w3 = w4 = w5 = 0

        g1 = g2 = g3 = g4 = 0

        a1 = a2 = 0

        v1 = v2 = v3 = v4 = v5 = 0

        for a in answer:

            if a.herramientas == 1:
                h1 += 1
            elif a.herramientas == 2:
                h2 += 1
            elif a.herramientas == 3:
                h3 += 1
            elif a.herramientas == 4:
                h4 += 1
            else:
                h5 += 1

            if a.colabora == 1:
                c1 += 1
                if a.tipoInvestigacion == 1:
                    d1 += 1
                elif a.tipoInvestigacion == 2:
                    d2 += 1
                elif a.tipoInvestigacion == 3:
                    d3 += 1
                elif a.tipoInvestigacion == 4:
                    d4 += 1
                else:
                    d5 += 1
            else:
                c2 += 1

            if a.participaRedes:
                r1 += 1
            else:
                r2 += 1

            if a.certificacion:
                z1 += 1
            else:
                z2 += 1

            if a.servicios == 1:
                s1 += 1
            elif a.servicios == 2:
                s2 += 1
            elif a.servicios == 3:
                s3 += 1
            elif a.servicios == 4:
                s4 += 1
            else:
                s5 += 1

            if a.idiomas == 1:
                i1 += 1
            elif a.idiomas == 2:
                i2 += 1
            elif a.idiomas == 3:
                i3 += 1
            elif a.idiomas == 4:
                i4 += 1
            else:
                i5 += 1

            if a.publicacion:
                p1 += 1
            else:
                p2 += 1

            if a.documentos == 1:
                w1 += 1
            elif a.documentos == 2:
                w2 += 1
            elif a.documentos == 3:
                w3 += 1
            elif a.documentos == 4:
                w4 += 1
            else:
                w5 += 1

            if a.calidad == 1:
                g1 += 1
            elif a.calidad == 2:
                g2 += 1
            elif a.calidad == 3:
                g3 += 1
            else:
                g4 += 1

            if a.asociacion:
                a1 += 1
            else:
                a2 += 1

            if a.etica == 1:
                v1 += 1
            elif a.etica == 2:
                v2 += 1
            elif a.etica == 3:
                v3 += 1
            elif a.etica == 4:
                v4 += 1
            else:
                v5 += 1

    return render(request, 'layouts/A4.html', {
        'anexo': True,
        'subtitle': 'DESEMPEÑO LABORAL',
        'utiliza': {'uno': h1, 'dos': h2, 'tres': h3, 'cuatro': h4, 'cinco': h5},
        'colabora': {'si': c1, 'no': c2},
        'tipo': {'uno': d1, 'dos': d2, 'tres': d3, 'cuatro': d4, 'cinco': d5},
        'redes': {'si': r1, 'no': r2},
        'certificacion': {'si': z1, 'no': z2},
        'servicio': {'uno': s1, 'dos': s2, 'tres': s3, 'cuatro': s4, 'cinco': s5},
        'idiomas': {'uno': i1, 'dos': i2, 'tres': i3, 'cuatro': i4, 'cinco': i5},
        'publicacion': {'si': p1, 'no': p2},
        'documentos': {'uno': w1, 'dos': w2, 'tres': w3, 'cuatro': w4, 'cinco': w5},
        'gestion': {'uno': g1, 'dos': g2, 'tres': g3, 'cuatro': g4},
        'asociacion': {'si': a1, 'no': a2},
        'etica': {'uno': v1, 'dos': v2, 'tres': v3, 'cuatro': v4, 'cinco': v5},
        'encuesta': p,
        'url': 'encuestaR/exportarA4'

    })


def exportarA4(request):
    if request.method == 'POST':
        p = request.POST.get('periodo')
        
        data = AnexoS4.objects.filter(
            folioEncuesta__in=Encuesta.objects.filter(lapso=p).values('folioEncuesta')
        ).select_related('folioEncuesta__curp').all()

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="anexoA4_export.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'CURP', 'Herramientas', 'Colabora en proyectos', 'Tipo de investigación',
            'Participa en redes', 'Tiene certificación', 'Servicios', 'Idiomas',
            'Publicación científica', 'Documentos', 'Gestión', 'Asociación profesional',
            'Ética'
        ])

        for a in data:
            egresado = a.folioEncuesta.curp

            writer.writerow([
                egresado.curp,
                a.get_herramientas_display() if a.herramientas else '',
                'Sí' if a.colabora == 1 else 'No',
                a.get_tipoInvestigacion_display() if a.tipoInvestigacion else '',
                'Sí' if a.participaRedes else 'No',
                'Sí' if a.certificacion else 'No',
                a.get_servicios_display() if a.servicios else '',
                a.get_idiomas_display() if a.idiomas else '',
                'Sí' if a.publicacion else 'No',
                a.get_documentos_display() if a.documentos else '',
                a.get_calidad_display() if a.calidad else '',
                'Sí' if a.asociacion else 'No',
                a.get_etica_display() if a.etica else ''
            ])

    return response


def A4(request):
    
    return render(request, 'layouts/A4.html', {
        'anexo': True,
        'subtitle': 'VALORES Y COMPETENCIAS'
    })


# exportaciones por usuario
def formulario_export_anexo(request):
    carreras = [
        "general", "Ing. Química", "Ing. Bioquímica"
    ]
    return render(request, 'exportarAnexoUsuario.html', {'carreras': carreras})


def exportar_anexo_por_carrera(request):
    carrera = request.GET.get('carrera')
    if not carrera:
        return HttpResponse("Carrera no seleccionada", status=400)

    rfc_admin = request.session.get('usuario_id')
    if not rfc_admin:
        return HttpResponse("Sesión no válida. Inicia sesión nuevamente.", status=401)

    try:
        admin = Administrador.objects.get(rfc=rfc_admin)
    except Administrador.DoesNotExist:
        return HttpResponse("Administrador no encontrado.", status=404)

    if admin.carrera.lower() != "general" and admin.carrera.lower() != carrera.lower():
        return HttpResponse("No tienes permiso para exportar esta carrera.", status=403)

    if carrera == "general":
        egresados_curps = list(Egresado.objects.values_list('curp', flat=True))
    else:
        egresados_curps = list(Egresado.objects.filter(
            carrera=carrera).values_list('curp', flat=True))

    ultimos = Encuesta.objects.filter(
        curp=OuterRef('folioEncuesta__curp')
    ).order_by('-lapso').values('folioEncuesta')[:1]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Anexo S1"
    ws2 = wb.create_sheet("Anexo S2")
    ws3 = wb.create_sheet("Anexo S3")
    ws4 = wb.create_sheet("Anexo S4")

    # 🟢 Hoja 1 - Anexo S1
    ws1.append(['CURP', 'Carrera', 'Lapso', 'Nombre completo', 'Correo',
               'Teléfono', 'Fecha Egreso', 'Titulado', 'Razón No Título', 'Otra razón'])

    anexos1 = AnexoS1.objects.filter(
        folioEncuesta_id=Subquery(ultimos),
        folioEncuesta__curp__curp__in=egresados_curps
    ).select_related('folioEncuesta__curp')

    for a in anexos1:
        ws1.append([
            a.folioEncuesta.curp.curp,
            a.folioEncuesta.curp.carrera or 'No definida',
            a.folioEncuesta.lapso,
            a.nombreCompleto or '',
            a.correo or '',
            a.telefono or '',
            a.fechaEgreso or '',
            a.get_titulado_display() if a.titulado is not None else '',
            a.get_razonNoTitulo_display() if a.razonNoTitulo else '',
            a.razonNoTituloOtra or ''
        ])


    # 🟢 Hoja 2 - Anexo S2
    ws2.append(['CURP', 'Carrera', 'Lapso', '¿Trabaja?', 'Razón no trabaja', 'Relación con carrera',
               'Antigüedad', 'Tiempo trabajo relacionado', 'Sector', 'Rol', 'Área', 'Medio trabajo', 'Satisfacción'])

    anexos2 = AnexoS2.objects.filter(
        folioEncuesta_id=Subquery(ultimos),
        folioEncuesta__curp__curp__in=egresados_curps
    ).select_related('folioEncuesta__curp')

    for a in anexos2:
        ws2.append([
            a.folioEncuesta.curp.curp,
            a.folioEncuesta.curp.carrera or 'No definida',
            a.folioEncuesta.lapso,
            a.get_trabaja_display() if a.trabaja is not None else '',
            a.get_razonNoTrabaja_display() if a.razonNoTrabaja else '',
            a.get_relacionTrabajoCarrera_display() if a.relacionTrabajoCarrera else '',
            a.get_antiguedad_display() if a.antiguedad else '',
            a.get_tiempoTrabajoRelacionado_display() if a.tiempoTrabajoRelacionado else '',
            a.get_sector_display() if a.sector else '',
            a.get_rol_display() if a.rol else '',
            a.get_area_display() if a.area else '',
            a.get_medioTrabajo_display() if a.medioTrabajo else '',
            a.get_satisfaccion_display() if a.satisfaccion else ''
        ])

    # 🟢 Hoja 3 - Anexo S3
    ws3.append(['CURP', 'Carrera', 'Lapso', 'Competencias', 'Satisfacción',
               'Educativo', '¿Contacto?', '¿Participa?', 'Aporte'])

    anexos3 = AnexoS3.objects.filter(
        folioEncuesta_id=Subquery(ultimos),
        folioEncuesta__curp__curp__in=egresados_curps
    ).select_related('folioEncuesta__curp')

    for a in anexos3:
        ws3.append([
            a.folioEncuesta.curp.curp,
            a.folioEncuesta.curp.carrera or 'No definida',
            a.folioEncuesta.lapso,
            a.get_competencias_display() if a.competencias else '',
            a.get_satisfaccion_display() if a.satisfaccion else '',
            a.get_educativo_display() if a.educativo else '',
            a.get_contacto_display() if a.contacto is not None else '',
            a.get_participar_display() if a.participar is not None else '',
            a.get_aporte_display() if a.aporte else ''
        ])

    # 🟢 Hoja 4 - Anexo S4
    ws4.append(['CURP', 'Carrera', 'Lapso', 'Herramientas', '¿Colabora?', 'Tipo investigación', '¿Participa en redes?',
               '¿Tiene certificación?', 'Servicios', 'Idiomas', '¿Publicación?', 'Documentos', 'Calidad', '¿Asociación?', 'Ética'])

    anexos4 = AnexoS4.objects.filter(
        folioEncuesta_id=Subquery(ultimos),
        folioEncuesta__curp__curp__in=egresados_curps
    ).select_related('folioEncuesta__curp')

    for a in anexos4:
        ws4.append([
            a.folioEncuesta.curp.curp,
            a.folioEncuesta.curp.carrera or 'No definida',
            a.folioEncuesta.lapso,
            a.get_herramientas_display() if a.herramientas else '',
            a.get_colabora_display() if a.colabora is not None else '',
            a.get_tipoInvestigacion_display() if a.tipoInvestigacion else '',
            a.get_participaRedes_display() if a.participaRedes is not None else '',
            a.get_certificacion_display() if a.certificacion is not None else '',
            a.get_servicios_display() if a.servicios else '',
            a.get_idiomas_display() if a.idiomas else '',
            a.get_publicacion_display() if a.publicacion is not None else '',
            a.get_documentos_display() if a.documentos else '',
            a.get_calidad_display() if a.calidad else '',
            a.get_asociacion_display() if a.asociacion is not None else '',
            a.get_etica_display() if a.etica else ''
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = 'anexo_general.xlsx' if carrera == "general" else f'anexo_{carrera}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)
    return response
